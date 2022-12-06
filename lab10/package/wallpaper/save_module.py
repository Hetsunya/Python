# -*- coding: utf-8 -*-
import openpyxl

def  data_Save(data_quantity, data_cost):
    wb = openpyxl.open("Result.xlsx", read_only=False)
    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['E1'] = "Wallpaper:"
    
    ws['E2'] = "Количество:"
    ws['E2'] = data_quantity

    ws['F3'] = "Общая цена:"
    ws['F3'].value = data_cost

    wb.save("Result.xlsx")

