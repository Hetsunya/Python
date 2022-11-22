# -*- coding: utf-8 -*-

from guizero import *
from openpyxl import Workbook
import lab10
app = App("Lab10")

def push_button():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    import datetime

    ws['A2'] = "Hello World!"

    # Save the file
    wb.save("sample.xlsx")
    """
    push_Button = PushButton(app, text='Сохранить данные')
    data_File = open('Data.doc', 'w+')
    data_File.write(data.value)
    """
