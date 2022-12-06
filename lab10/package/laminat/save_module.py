# -*- coding: utf-8 -*-
import openpyxl

def  L_data_Save(data_quantity, data_cost):
    wb = openpyxl.open("Result.xlsx", read_only=False)

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 'Laminat'
    ws['A2'] = "Количество:"
    ws['B2'] = data_quantity

    ws['A3'] = "Общая цена:"
    ws['B3'].value = data_cost

    wb.save("Result.xlsx")

