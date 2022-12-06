# -*- coding: utf-8 -*-
import openpyxl

def  data_Save(data_quantity, data_cost):
    wb = openpyxl.open("Result.xlsx", read_only=False)

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['C1'] = 'Tile'
    ws['C2'] = "Количество:"
    ws['C2'] = data_quantity

    ws['D3'] = "Общая цена:"
    ws['D3'].value = data_cost

    wb.save("Result.xlsx")

